"""Inventory API endpoints."""
import logging
from typing import List, Optional, Any
from fastapi import APIRouter, Depends, File, HTTPException, status, Request, Query, UploadFile
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user, get_inventory_manager_or_admin
from app.models.user import User
from app.schemas.base import InventoryUpdate, InventoryResponse
from app.services.inventory_service import InventoryService

router = APIRouter()
logger = logging.getLogger(__name__)


def get_client_ip(request: Request) -> str:
    """Extract client IP address from request."""
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


@router.get("/", response_model=List[InventoryResponse])
def list_inventory(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=5000),
    supply_id: Optional[int] = Query(None, description="Filter by supply ID"),
    location: Optional[str] = Query(None, description="Filter by location"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """
    List all inventory items.
    
    Supports filtering by supply_id and location.
    All authenticated users can access this endpoint.
    """
    inventory_service = InventoryService(db)
    items = inventory_service.get_inventory_items(
        skip=skip,
        limit=limit,
        supply_id=supply_id,
        location=location
    )
    return items


@router.get("/low-stock", response_model=List[InventoryResponse])
def get_low_stock_items(
    threshold: float = Query(1.0, ge=0.1, le=2.0, description="Threshold multiplier for safety stock"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """
    Get inventory items with low stock.
    
    Returns items where current_stock <= safety_stock * threshold.
    Default threshold is 1.0 (at or below safety stock).
    """
    inventory_service = InventoryService(db)
    items = inventory_service.get_low_stock_items(threshold_multiplier=threshold)
    return items


@router.get("/expiring", response_model=List[InventoryResponse])
def get_expiring_items(
    days: int = Query(30, ge=1, le=365, description="Number of days to look ahead"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """
    Get inventory items expiring within the specified number of days.
    
    Default is 30 days.
    """
    inventory_service = InventoryService(db)
    items = inventory_service.get_expiring_items(days_threshold=days)
    return items


# ── Smart Medical spec 6 — Static path routes phải khai báo trước "/{inventory_id}" ──


@router.get("/template", include_in_schema=True)
async def download_inventory_template():
    """File CSV mẫu cho Import tồn kho đầu kỳ (4 bệnh hô hấp + 15 thuốc/vật tư)."""
    from fastapi.responses import StreamingResponse

    sample = (
        "supply_code,drug_code,ten_hoat_chat,unit,group_name,category,current_stock,safety_stock,expiry_date,lead_time_days\n"
        "VT001,A2A210200000133,Paracetamol,Viên,Thuốc hạ sốt giảm đau,medicine,8500,2000,2027-06-30,7\n"
        "VT002,A2C816100000074,Natri clorid,Chai,Dung dịch/dịch truyền,medicine,420,500,2026-12-31,10\n"
        "VT003,A2C715800000009,N-acetylcysteine,Gói,Thuốc long đờm,medicine,3200,1000,2027-09-30,7\n"
        "VT004,A2A300000000058,Fexofenadin,Viên,Kháng histamin,medicine,2400,500,2027-03-31,7\n"
        "VT005,A2A610720000036,Amoxicilin + acid clavulanic,Viên,Kháng sinh uống,medicine,2400,1500,2026-12-31,7\n"
    )
    return StreamingResponse(
        iter([sample]),
        media_type="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=inventory_template.csv"
        },
    )


@router.post("/import")
async def _import_inventory_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_inventory_manager_or_admin),
) -> Any:
    """Import tồn kho đầu kỳ từ file CSV (spec 6.5)."""
    return await _do_inventory_import(file, db, current_user)


@router.get("/export")
async def _export_inventory_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Any:
    """Xuất báo cáo tồn kho ra file Excel (.xlsx) — spec 6.6."""
    return _do_inventory_export(db)


@router.post("/sync-safety-stock")
async def sync_safety_stock_from_forecast(
    forecast_month: Optional[str] = Query(
        None, description="Tháng dự báo (YYYY-MM-DD), để trống = tháng gần nhất"
    ),
    buffer_rate: float = Query(15.0, ge=0, le=100, description="% dự phòng"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_inventory_manager_or_admin),
) -> Any:
    """
    Cập nhật ngưỡng an toàn (safety_stock) trong inventory từ kết quả dự báo.
    
    Logic:
    1. Lấy kết quả dự báo từ supply_recommendations cho tháng chỉ định
    2. Tính calculated_safety_stock = need_before_buffer × (1 + buffer_rate/100)
    3. Cập nhật inventory.safety_stock cho các vật tư tương ứng
    
    Returns:
        - updated: số vật tư đã cập nhật
        - skipped: số vật tư bỏ qua (không có dự báo)
    """
    from datetime import datetime as _dt
    from app.models.supply_recommendation import SupplyRecommendation

    # Parse forecast_month
    if forecast_month:
        try:
            target_month = _dt.fromisoformat(forecast_month).date()
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail="forecast_month phải có định dạng YYYY-MM-DD",
            )
    else:
        # Lấy tháng gần nhất có dữ liệu
        latest = (
            db.query(SupplyRecommendation.forecast_month)
            .order_by(SupplyRecommendation.forecast_month.desc())
            .first()
        )
        if not latest:
            raise HTTPException(
                status_code=404,
                detail="Không tìm thấy dữ liệu dự báo nào. Vui lòng chạy phân tích dự báo trước.",
            )
        target_month = latest[0]

    # Lấy tất cả recommendations cho tháng đó
    recommendations = (
        db.query(SupplyRecommendation)
        .filter(SupplyRecommendation.forecast_month == target_month)
        .all()
    )

    if not recommendations:
        raise HTTPException(
            status_code=404,
            detail=f"Không tìm thấy dữ liệu dự báo cho tháng {target_month}",
        )

    # Group theo supply_id và tính tổng need_before_buffer
    supply_needs: dict[int, float] = {}
    for rec in recommendations:
        sid = rec.supply_id
        if sid not in supply_needs:
            supply_needs[sid] = 0
        supply_needs[sid] += rec.need_before_buffer

    # Cập nhật inventory.safety_stock
    updated = 0
    skipped = 0
    from app.models.inventory import Inventory as InventoryModel

    for supply_id, need_before_buffer in supply_needs.items():
        # Tính ngưỡng AT theo công thức
        calculated_safety = round(need_before_buffer * (1 + buffer_rate / 100))

        # Tìm inventory record
        inv = (
            db.query(InventoryModel)
            .filter(InventoryModel.supply_id == supply_id)
            .first()
        )

        if inv:
            inv.safety_stock = calculated_safety
            updated += 1
        else:
            skipped += 1

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error(f"Failed to sync safety stock: {exc}")
        raise HTTPException(
            status_code=500,
            detail="Không thể cập nhật ngưỡng an toàn",
        )

    logger.info(
        f"Synced safety_stock from forecast {target_month}: "
        f"updated={updated}, skipped={skipped} by user={current_user.username}"
    )

    return {
        "forecast_month": target_month.isoformat(),
        "buffer_rate": buffer_rate,
        "updated": updated,
        "skipped": skipped,
        "message": f"Đã cập nhật ngưỡng an toàn cho {updated} vật tư từ dự báo tháng {target_month.strftime('%m/%Y')}",
    }


@router.get("/{inventory_id}", response_model=InventoryResponse)
def get_inventory_item(
    inventory_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    """
    Get inventory item by ID.
    
    All authenticated users can access this endpoint.
    """
    inventory_service = InventoryService(db)
    item = inventory_service.get_inventory_by_id(inventory_id)
    return item


@router.delete("/{inventory_id}", status_code=status.HTTP_200_OK)
def delete_inventory(
    inventory_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_inventory_manager_or_admin),
) -> Any:
    """
    Delete an inventory record.

    Only Inventory Managers and Administrators can delete inventory entries.
    Resolves any open alerts attached to the same supply afterwards.
    """
    from app.models.inventory import Inventory as InventoryModel
    from app.models.alert import Alert as AlertModel
    from datetime import datetime, timezone

    item = (
        db.query(InventoryModel)
        .filter(InventoryModel.id == inventory_id)
        .first()
    )
    if not item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Inventory {inventory_id} not found",
        )

    supply_id = item.supply_id
    db.delete(item)

    # Resolve open alerts tied to this supply if no inventory rows remain
    remaining = (
        db.query(InventoryModel)
        .filter(
            InventoryModel.supply_id == supply_id,
            InventoryModel.id != inventory_id,
        )
        .count()
    )
    if remaining == 0:
        now = datetime.now(timezone.utc)
        for alert in (
            db.query(AlertModel)
            .filter(
                AlertModel.supply_id == supply_id,
                AlertModel.is_resolved == False,  # noqa: E712
            )
            .all()
        ):
            alert.is_resolved = True
            alert.resolved_at = now

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error(f"Failed to delete inventory {inventory_id}: {exc}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete inventory record",
        )

    logger.info(
        f"Inventory {inventory_id} (supply_id={supply_id}) deleted by "
        f"user={current_user.username}"
    )
    return {"message": "Inventory deleted", "id": inventory_id}


@router.put("/{inventory_id}", response_model=InventoryResponse)
def update_inventory(
    inventory_id: int,
    inventory_data: InventoryUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_inventory_manager_or_admin)
) -> Any:
    """
    Update inventory stock levels.
    
    Only Inventory Managers and Administrators can update inventory.
    """
    inventory_service = InventoryService(db)
    client_ip = get_client_ip(request)
    
    try:
        updated_item = inventory_service.update_inventory(
            inventory_id=inventory_id,
            inventory_data=inventory_data,
            updated_by_user_id=current_user.id,
            ip_address=client_ip
        )
        return updated_item
    except Exception as e:
        logger.error(f"Failed to update inventory {inventory_id}: {str(e)}")
        raise


@router.post("/batch-update", response_model=List[InventoryResponse])
def batch_update_inventory(
    updates: List[dict],
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_inventory_manager_or_admin)
) -> Any:
    """
    Batch update multiple inventory items.
    
    Request body should be a list of objects with:
    - inventory_id (required): ID of inventory item to update
    - current_stock (optional): New current stock value
    - safety_stock (optional): New safety stock value
    
    Only Inventory Managers and Administrators can perform batch updates.
    
    Example:
    [
        {"inventory_id": 1, "current_stock": 500},
        {"inventory_id": 2, "current_stock": 300, "safety_stock": 100}
    ]
    """
    inventory_service = InventoryService(db)
    client_ip = get_client_ip(request)
    
    try:
        updated_items = inventory_service.batch_update_inventory(
            updates=updates,
            updated_by_user_id=current_user.id,
            ip_address=client_ip
        )
        return updated_items
    except Exception as e:
        logger.error(f"Failed to batch update inventory: {str(e)}")
        raise



# ── Smart Medical spec 6.5: Helpers ─────────────────────────────────────────


async def _do_inventory_import(file: "UploadFile", db: Session, current_user: User) -> Any:
    """Logic import tồn kho đầu kỳ (spec 6.5)."""
    import csv as _csv
    import io as _io
    from datetime import datetime as _dt

    from app.models.medical_supply import MedicalSupply
    from app.models.inventory import Inventory as InventoryModel

    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="File must be a CSV")

    raw = await file.read()
    text = raw.decode("utf-8", errors="replace")
    reader = list(_csv.DictReader(_io.StringIO(text)))

    if not reader:
        return {"imported": 0, "updated": 0, "skipped": 0, "errors": []}

    imported = 0
    updated = 0
    skipped = 0
    errors: list[dict] = []

    for idx, row in enumerate(reader, start=2):
        # Hỗ trợ format mới (supply_code/drug_code/ten_hoat_chat) và format cũ (supply_name)
        supply_code = (row.get("supply_code") or "").strip()
        drug_code = (row.get("drug_code") or "").strip()
        ten_hoat_chat = (row.get("ten_hoat_chat") or row.get("supply_name") or "").strip()
        category = (row.get("category") or "other").strip()
        unit = (row.get("unit") or "Cái").strip()
        group_name = (row.get("group_name") or category or "Khác").strip()

        if not ten_hoat_chat and not supply_code and not drug_code:
            skipped += 1
            errors.append({"row": idx, "reason": "Thiếu supply_code / drug_code / ten_hoat_chat"})
            continue

        # Parse số
        try:
            # Hỗ trợ cả current_stock và stock_quantity
            cur_stock = int(float((row.get("current_stock") or row.get("stock_quantity") or "0").strip() or 0))
            saf_stock = int(float((row.get("safety_stock") or "0").strip() or 0))
        except ValueError:
            skipped += 1
            errors.append(
                {"row": idx, "reason": "current_stock / safety_stock không phải số"}
            )
            continue

        if cur_stock < 0 or saf_stock < 0:
            skipped += 1
            errors.append({"row": idx, "reason": "Tồn kho / ngưỡng AT không được âm"})
            continue

        try:
            lead_time = int(
                float((row.get("lead_time_days") or "").strip() or 0)
            )
        except ValueError:
            lead_time = None

        expiry_str = (row.get("expiry_date") or "").strip()
        expiry_date = None
        if expiry_str:
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    expiry_date = _dt.strptime(expiry_str, fmt).date()
                    break
                except ValueError:
                    continue

        # 1. Tìm MedicalSupply theo supply_code, drug_code, hoặc ten_hoat_chat
        supply = None
        if supply_code:
            supply = (
                db.query(MedicalSupply)
                .filter(MedicalSupply.supply_code == supply_code)
                .first()
            )
        if supply is None and drug_code:
            supply = (
                db.query(MedicalSupply)
                .filter(MedicalSupply.drug_code == drug_code)
                .first()
            )
        if supply is None and ten_hoat_chat:
            supply = (
                db.query(MedicalSupply)
                .filter(MedicalSupply.ten_hoat_chat == ten_hoat_chat)
                .first()
            )

        is_new_supply = supply is None
        if is_new_supply:
            # Tạo mới với các trường bắt buộc
            if not supply_code:
                supply_code = f"VT_AUTO_{int(_dt.now().timestamp())}_{idx}"
            if not drug_code:
                drug_code = supply_code
            if not ten_hoat_chat:
                ten_hoat_chat = supply_code
            supply = MedicalSupply(
                supply_code=supply_code,
                drug_code=drug_code,
                ten_hoat_chat=ten_hoat_chat,
                unit=unit,
                group_name=group_name,
                category=category,
                lead_time_days=lead_time,
            )
            db.add(supply)
            db.flush()
        else:
            # Cập nhật metadata nếu có thay đổi
            if category and category != "other":
                supply.category = category
            if unit:
                supply.unit = unit
            if group_name and group_name != "Khác":
                supply.group_name = group_name
            if lead_time is not None:
                supply.lead_time_days = lead_time

        # 2. Tìm/Tạo Inventory record
        inv = (
            db.query(InventoryModel)
            .filter(InventoryModel.supply_id == supply.id)
            .first()
        )
        if inv is None:
            db.add(
                InventoryModel(
                    supply_id=supply.id,
                    current_stock=cur_stock,
                    safety_stock=saf_stock,
                    expiry_date=expiry_date,
                )
            )
            imported += 1
        else:
            inv.current_stock = cur_stock
            inv.safety_stock = saf_stock
            if expiry_date:
                inv.expiry_date = expiry_date
            updated += 1

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to save: {exc}")

    logger.info(
        "Inventory import by user=%s: imported=%d updated=%d skipped=%d",
        current_user.username, imported, updated, skipped,
    )
    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:200],
        "errors_truncated": len(errors) > 200,
    }


def _do_inventory_export(db: Session) -> Any:
    """Logic xuất Excel báo cáo tồn kho."""
    import io as _io
    from datetime import datetime as _dt
    from fastapi.responses import Response

    from app.models.medical_supply import MedicalSupply
    from app.models.inventory import Inventory as InventoryModel

    rows = (
        db.query(
            InventoryModel.id,
            InventoryModel.current_stock,
            InventoryModel.safety_stock,
            InventoryModel.expiry_date,
            MedicalSupply.id.label("supply_id"),
            MedicalSupply.supply_code,
            MedicalSupply.drug_code,
            MedicalSupply.ten_hoat_chat,
            MedicalSupply.group_name,
            MedicalSupply.category,
            MedicalSupply.unit,
            MedicalSupply.lead_time_days,
        )
        .join(MedicalSupply, MedicalSupply.id == InventoryModel.supply_id)
        .order_by(MedicalSupply.supply_code)
        .all()
    )

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed — không thể xuất Excel",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tồn kho"

    headers = [
        "Mã VT",
        "Tên vật tư",
        "Loại",
        "ĐVT",
        "Tồn kho",
        "Ngưỡng AT",
        "Trạng thái",
        "Hạn dùng",
        "Lead time (ngày)",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for col_num, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_num, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center

    category_label = {
        "medicine": "Thuốc",
        "mask": "Khẩu trang",
        "glove": "Găng tay",
        "test_kit": "Kit XN",
        "disinfectant": "Hoá chất",
        "iv_fluid": "Dịch truyền",
        "other": "Khác",
    }
    prefix_map = {
        "medicine": "VT",
        "mask": "TB",
        "glove": "TB",
        "test_kit": "TB",
        "disinfectant": "HC",
        "iv_fluid": "VT",
        "other": "VT",
    }
    status_color = {
        "Bình thường": "16A34A",
        "Dưới ngưỡng": "D97706",
        "Cần nhập gấp": "DC2626",
    }

    for r_idx, r in enumerate(rows, start=2):
        cur = r.current_stock or 0
        saf = r.safety_stock or 0
        if cur <= 0:
            status_text = "Cần nhập gấp"
        elif saf > 0 and cur < saf * 0.3:
            status_text = "Cần nhập gấp"
        elif cur <= saf:
            status_text = "Dưới ngưỡng"
        else:
            status_text = "Bình thường"

        # Sử dụng supply_code mới (VT001-VT015), không cần prefix
        code = r.supply_code or f"VT-{str(r.supply_id).zfill(4)}"

        ws.cell(row=r_idx, column=1, value=code)
        ws.cell(row=r_idx, column=2, value=r.ten_hoat_chat)
        ws.cell(row=r_idx, column=3, value=r.group_name or category_label.get(r.category or "", r.category))
        ws.cell(row=r_idx, column=4, value=r.unit)
        ws.cell(row=r_idx, column=5, value=cur)
        ws.cell(row=r_idx, column=6, value=saf)
        status_cell = ws.cell(row=r_idx, column=7, value=status_text)
        status_cell.font = Font(color=status_color.get(status_text, "000000"), bold=True)
        ws.cell(
            row=r_idx,
            column=8,
            value=r.expiry_date.strftime("%d/%m/%Y") if r.expiry_date else "",
        )
        ws.cell(row=r_idx, column=9, value=r.lead_time_days or "")

    widths = [12, 38, 14, 8, 10, 10, 16, 12, 14]
    for col_num, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"inventory_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
