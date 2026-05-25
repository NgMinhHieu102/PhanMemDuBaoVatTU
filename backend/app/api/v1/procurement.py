"""
Procurement Planning API endpoints.

Provides full CRUD access to procurement plans, plan generation via the
ProcurementPlanner, an approval workflow, and PDF/Excel export.

Routes
------
GET    /api/v1/procurement              – List all procurement plans
POST   /api/v1/procurement/generate    – Auto-generate plans from critical/high alerts
GET    /api/v1/procurement/export      – Export plans to PDF or Excel
GET    /api/v1/procurement/{id}        – Get plan by ID
PUT    /api/v1/procurement/{id}        – Update a plan
DELETE /api/v1/procurement/{id}        – Delete a plan
POST   /api/v1/procurement/{id}/approve – Approve a plan
"""

import io
import logging
from datetime import date, datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user, get_admin_user
from app.models.procurement_plan import ProcurementPlan
from app.models.medical_supply import MedicalSupply
from app.models.user import User
from app.procurement.procurement_planner import ProcurementPlanner
from app.schemas.base import (
    ProcurementPlanCreate,
    ProcurementPlanUpdate,
    ProcurementPlanResponse,
    ProcurementGenerateRequest,
    ProcurementGenerateResponse,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["procurement"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _enrich_response(plan: ProcurementPlan) -> dict:
    """Attach supply_name to plan dict so ProcurementPlanResponse can be populated."""
    return {
        "id": plan.id,
        "supply_id": plan.supply_id,
        "supply_name": plan.supply.name if plan.supply else None,
        "order_quantity": plan.order_quantity,
        "order_date": plan.order_date,
        "expected_delivery_date": plan.expected_delivery_date,
        "estimated_cost": float(plan.estimated_cost) if plan.estimated_cost is not None else None,
        "priority": plan.priority,
        "status": plan.status,
        "notes": plan.notes,
        "created_at": plan.created_at,
    }


def _get_plan_or_404(plan_id: int, db: Session) -> ProcurementPlan:
    """Fetch a plan by ID or raise 404."""
    plan = (
        db.query(ProcurementPlan)
        .filter(ProcurementPlan.id == plan_id)
        .first()
    )
    if not plan:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Procurement plan with ID {plan_id} not found",
        )
    return plan


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("", response_model=List[ProcurementPlanResponse])
async def list_procurement_plans(
    status_filter: Optional[str] = Query(
        None,
        alias="status",
        description="Filter by status: pending, approved, cancelled",
    ),
    priority: Optional[str] = Query(
        None,
        description="Filter by priority: critical, high, normal",
    ),
    supply_id: Optional[int] = Query(None, description="Filter by supply ID"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum number of results"),
    offset: int = Query(0, ge=0, description="Number of results to skip"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[ProcurementPlanResponse]:
    """
    List all procurement plans with optional filters.

    Supports filtering by `status`, `priority`, and `supply_id`.
    Results are paginated via `limit` / `offset`.
    """
    logger.info(
        "List procurement plans requested by user=%s "
        "status=%s priority=%s supply_id=%s",
        current_user.username, status_filter, priority, supply_id,
    )
    query = db.query(ProcurementPlan)

    if status_filter:
        query = query.filter(ProcurementPlan.status == status_filter)
    if priority:
        query = query.filter(ProcurementPlan.priority == priority)
    if supply_id:
        query = query.filter(ProcurementPlan.supply_id == supply_id)

    # Sort: critical first, then high, then normal; newest first within same priority
    priority_order_map = {"critical": 0, "high": 1, "normal": 2}
    plans = (
        query
        .order_by(ProcurementPlan.order_date.asc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    # Re-sort in Python to apply priority ordering
    plans.sort(key=lambda p: priority_order_map.get(p.priority or "normal", 2))

    return [ProcurementPlanResponse(**_enrich_response(p)) for p in plans]


@router.post("/generate", response_model=ProcurementGenerateResponse, status_code=status.HTTP_201_CREATED)
async def generate_procurement_plans(
    request: ProcurementGenerateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ProcurementGenerateResponse:
    """
    Auto-generate procurement plans from critical and high severity alerts.

    Uses the ProcurementPlanner to calculate optimal order quantities, timing,
    and costs.  Existing pending plans for the same supply are replaced.
    Critical supplies are prioritised over high-priority ones.

    Args:
        request.forecast_days: How many days ahead to plan for (7–90, default 30).
    """
    logger.info(
        "Generate procurement plans requested by user=%s forecast_days=%d",
        current_user.username, request.forecast_days,
    )
    try:
        planner = ProcurementPlanner(db)
        plan_items = planner.generate_plan(forecast_days=request.forecast_days)

        if not plan_items:
            return ProcurementGenerateResponse(
                message="No procurement plans generated – no supply requirements found",
                plans_generated=0,
                critical_plans=0,
                high_plans=0,
                normal_plans=0,
                plans=[],
            )

        saved_plans = planner.save_plan(plan_items, created_by=current_user.id)

        enriched = [ProcurementPlanResponse(**_enrich_response(p)) for p in saved_plans]

        critical = sum(1 for p in saved_plans if p.priority == "critical")
        high = sum(1 for p in saved_plans if p.priority == "high")
        normal = sum(1 for p in saved_plans if p.priority == "normal")

        logger.info(
            "Generated %d procurement plans (critical=%d, high=%d, normal=%d) for user=%s",
            len(saved_plans), critical, high, normal, current_user.username,
        )

        return ProcurementGenerateResponse(
            message=f"Successfully generated {len(saved_plans)} procurement plan(s)",
            plans_generated=len(saved_plans),
            critical_plans=critical,
            high_plans=high,
            normal_plans=normal,
            plans=enriched,
        )

    except Exception as exc:
        logger.error("Error generating procurement plans: %s", exc, exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate procurement plans",
        )


@router.get("/export")
async def export_procurement_plans(
    format: str = Query(
        "excel",
        description="Export format: pdf or excel",
        pattern="^(pdf|excel)$",
    ),
    status_filter: Optional[str] = Query(None, alias="status", description="Filter by status"),
    priority: Optional[str] = Query(None, description="Filter by priority"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Response:
    """
    Export procurement plans to PDF or Excel.

    Columns: Supply Name, Order Quantity, Order Date, Expected Delivery,
    Estimated Cost, Priority, Status.

    Query params:
    - format: ``pdf`` or ``excel`` (default ``excel``)
    - status: optional status filter
    - priority: optional priority filter
    """
    logger.info(
        "Export procurement plans format=%s requested by user=%s",
        format, current_user.username,
    )

    # Fetch plans with optional filters
    query = db.query(ProcurementPlan)
    if status_filter:
        query = query.filter(ProcurementPlan.status == status_filter)
    if priority:
        query = query.filter(ProcurementPlan.priority == priority)

    plans = query.order_by(ProcurementPlan.order_date.asc()).all()

    # Build rows for export
    rows = []
    for plan in plans:
        supply_name = plan.supply.name if plan.supply else f"Supply #{plan.supply_id}"
        rows.append({
            "supply_name": supply_name,
            "order_quantity": plan.order_quantity,
            "order_date": str(plan.order_date) if plan.order_date else "",
            "expected_delivery_date": str(plan.expected_delivery_date) if plan.expected_delivery_date else "",
            "estimated_cost": float(plan.estimated_cost) if plan.estimated_cost else 0.0,
            "priority": plan.priority or "",
            "status": plan.status or "",
        })

    if format == "excel":
        return _export_excel(rows)
    else:
        return _export_pdf(rows)


def _export_excel(rows: list) -> Response:
    """Generate an Excel file from procurement plan rows."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library not available for Excel export",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kế hoạch nhập kho"

    # Header row (Vietnamese)
    headers = [
        "Tên vật tư",
        "SL nhập",
        "Ngày đặt",
        "Ngày giao",
        "Chi phí (VND)",
        "Mức ưu tiên",
        "Trạng thái",
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Map priority/status sang VN
    priority_label = {"critical": "Khẩn cấp", "high": "Cao", "normal": "Thường"}
    status_label = {
        "pending": "Chờ duyệt",
        "approved": "Đã duyệt",
        "cancelled": "Đã huỷ",
        "ordered": "Đã đặt",
    }
    priority_colours = {"critical": "DC2626", "high": "EA580C", "normal": "16A34A"}

    # Data rows
    for row_num, row in enumerate(rows, 2):
        prio_key = (row["priority"] or "").lower()
        status_key = (row["status"] or "").lower()

        ws.cell(row=row_num, column=1, value=row["supply_name"])
        ws.cell(row=row_num, column=2, value=row["order_quantity"])
        ws.cell(row=row_num, column=3, value=row["order_date"])
        ws.cell(row=row_num, column=4, value=row["expected_delivery_date"])
        ws.cell(row=row_num, column=5, value=row["estimated_cost"])
        priority_cell = ws.cell(
            row=row_num, column=6, value=priority_label.get(prio_key, prio_key.capitalize())
        )
        ws.cell(
            row=row_num, column=7, value=status_label.get(status_key, status_key.capitalize())
        )

        # Colour-code priority
        prio_colour = priority_colours.get(prio_key, "000000")
        priority_cell.font = Font(color=prio_colour, bold=True)

    # Auto-fit columns
    widths = [38, 10, 12, 12, 16, 14, 14]
    for col_num, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = w

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote as _quote
    from app.api.v1.reports import _ascii_filename

    name = f"ke_hoach_nhap_kho_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_ascii_filename(name)}"; '
                f"filename*=UTF-8''{_quote(name)}"
            )
        },
    )


def _export_pdf(rows: list) -> Response:
    """Generate a PDF file from procurement plan rows."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="reportlab library not available for PDF export",
        )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    # Đăng ký font Unicode (DejaVu) cho tiếng Việt — dùng chung với reports module
    from app.api.v1 import reports as _reports

    _reports._register_unicode_fonts()
    PDF_FONT_BOLD = _reports.PDF_FONT_BOLD
    PDF_FONT_REGULAR = _reports.PDF_FONT_REGULAR

    styles = getSampleStyleSheet()
    _reports._patch_styles_for_unicode(styles)
    story = []

    # Title (Vietnamese)
    title = Paragraph(
        f"Báo cáo Kế hoạch Nhập kho - tạo lúc {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["Title"],
    )
    story.append(title)
    story.append(Spacer(1, 0.4 * cm))

    # Cell style cho text dài (tự wrap)
    from reportlab.lib.styles import ParagraphStyle
    cell_style = ParagraphStyle(
        "PlanCell",
        parent=styles["Normal"],
        fontName=PDF_FONT_REGULAR,
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )

    # Map priority + status sang tiếng Việt
    priority_label = {
        "critical": "Khẩn cấp",
        "high": "Cao",
        "normal": "Thường",
        "": "—",
    }
    status_label = {
        "pending": "Chờ duyệt",
        "approved": "Đã duyệt",
        "cancelled": "Đã huỷ",
        "ordered": "Đã đặt",
    }

    # Table header (Vietnamese)
    table_data = [
        ["Tên vật tư", "SL nhập", "Ngày đặt", "Ngày giao", "Chi phí (VND)", "Mức ưu tiên", "Trạng thái"],
    ]

    for row in rows:
        prio_key = (row["priority"] or "").lower()
        status_key = (row["status"] or "").lower()
        table_data.append([
            Paragraph(row["supply_name"], cell_style),
            str(row["order_quantity"]),
            row["order_date"],
            row["expected_delivery_date"],
            f"{row['estimated_cost']:,.0f}",
            priority_label.get(prio_key, prio_key.capitalize() if prio_key else "—"),
            status_label.get(status_key, status_key.capitalize() if status_key else "—"),
        ])

    col_widths = [6 * cm, 2 * cm, 3 * cm, 3 * cm, 3.5 * cm, 2.5 * cm, 3 * cm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    priority_colour_map = {
        "critical": colors.red,
        "high": colors.orange,
        "normal": colors.green,
        "": colors.black,
    }

    # Base style
    table_style = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Data rows
        ("FONTNAME", (0, 1), (-1, -1), PDF_FONT_REGULAR),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    # Colour-code priority column (col 5)
    for row_num, row in enumerate(rows, 1):
        prio = row["priority"].lower()
        prio_colour = priority_colour_map.get(prio, colors.black)
        table_style.append(("TEXTCOLOR", (5, row_num), (5, row_num), prio_colour))
        table_style.append(("FONTNAME", (5, row_num), (5, row_num), PDF_FONT_BOLD))

    table.setStyle(TableStyle(table_style))
    story.append(table)

    doc.build(story)
    buf.seek(0)

    from urllib.parse import quote as _quote
    from app.api.v1.reports import _ascii_filename

    name = f"ke_hoach_nhap_kho_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_ascii_filename(name)}"; '
                f"filename*=UTF-8''{_quote(name)}"
            )
        },
    )


@router.get("/{plan_id}", response_model=ProcurementPlanResponse)
async def get_procurement_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ProcurementPlanResponse:
    """
    Retrieve a single procurement plan by its ID.

    Returns 404 if the plan does not exist.
    """
    logger.info(
        "Get procurement plan id=%d requested by user=%s",
        plan_id, current_user.username,
    )
    plan = _get_plan_or_404(plan_id, db)
    return ProcurementPlanResponse(**_enrich_response(plan))


@router.put("/{plan_id}", response_model=ProcurementPlanResponse)
async def update_procurement_plan(
    plan_id: int,
    payload: ProcurementPlanUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ProcurementPlanResponse:
    """
    Update a procurement plan.

    Only non-None fields in the request body are updated.
    Approved plans cannot be modified.
    """
    logger.info(
        "Update procurement plan id=%d requested by user=%s",
        plan_id, current_user.username,
    )
    plan = _get_plan_or_404(plan_id, db)

    if plan.status == "approved":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Approved plans cannot be modified",
        )

    update_data = payload.model_dump(exclude_none=True)
    for field, value in update_data.items():
        setattr(plan, field, value)

    try:
        db.commit()
        db.refresh(plan)
    except Exception as exc:
        logger.error("Error updating procurement plan id=%d: %s", plan_id, exc)
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update procurement plan",
        )

    logger.info("Updated procurement plan id=%d by user=%s", plan_id, current_user.username)
    return ProcurementPlanResponse(**_enrich_response(plan))


@router.delete("/{plan_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_procurement_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> None:
    """
    Delete a procurement plan.

    Approved plans cannot be deleted.
    Returns 204 No Content on success.
    """
    logger.info(
        "Delete procurement plan id=%d requested by user=%s",
        plan_id, current_user.username,
    )
    plan = _get_plan_or_404(plan_id, db)

    if plan.status == "approved":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Approved plans cannot be deleted",
        )

    try:
        db.delete(plan)
        db.commit()
    except Exception as exc:
        logger.error("Error deleting procurement plan id=%d: %s", plan_id, exc)
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete procurement plan",
        )

    logger.info("Deleted procurement plan id=%d by user=%s", plan_id, current_user.username)


@router.post("/{plan_id}/approve", response_model=ProcurementPlanResponse)
async def approve_procurement_plan(
    plan_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ProcurementPlanResponse:
    """
    Approve a procurement plan.

    Sets `status` to ``approved``.
    Returns 400 if the plan is already approved.
    Returns 404 if the plan does not exist.
    """
    logger.info(
        "Approve procurement plan id=%d requested by user=%s",
        plan_id, current_user.username,
    )
    plan = _get_plan_or_404(plan_id, db)

    if plan.status == "approved":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Plan is already approved",
        )

    plan.status = "approved"

    try:
        db.commit()
        db.refresh(plan)
    except Exception as exc:
        logger.error("Error approving procurement plan id=%d: %s", plan_id, exc)
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to approve procurement plan",
        )

    logger.info(
        "Approved procurement plan id=%d by user=%s", plan_id, current_user.username
    )
    return ProcurementPlanResponse(**_enrich_response(plan))
