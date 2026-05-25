"""
Reports API endpoints.

Provides analytical reports for the MedForecast AI system, including
consumption analytics, forecast accuracy over time, inventory turnover,
and PDF export functionality.

All report endpoints support date range and location filtering.

Routes
------
GET  /api/v1/reports/consumption         – Consumption report by supply category
GET  /api/v1/reports/forecast-accuracy   – Forecast accuracy metrics over time
GET  /api/v1/reports/inventory-turnover  – Inventory turnover rates
POST /api/v1/reports/export              – Export any report to PDF
"""

import io
import logging
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user
from app.models.disease_forecast import DiseaseForecast
from app.models.inventory import Inventory
from app.models.medical_supply import MedicalSupply
from app.models.supply_requirement import SupplyRequirement
from app.models.user import User

logger = logging.getLogger(__name__)

router = APIRouter(tags=["reports"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _default_date_range(
    start_date: Optional[date],
    end_date: Optional[date],
    default_days_back: int = 30,
    default_days_forward: int = 60,
):
    """Return (start, end) dates spanning past consumption + future forecast.

    Mặc định mở rộng đến ngày hiện tại + 60 ngày để bao gồm cả các yêu cầu vật
    tư đã được dự báo cho tương lai (supply_requirements / disease_forecasts).
    """
    today = date.today()
    end = end_date or (today + timedelta(days=default_days_forward))
    start = start_date or (today - timedelta(days=default_days_back))
    return start, end


# ── Consumption Report ────────────────────────────────────────────────────────

@router.get("/consumption")
async def get_consumption_report(
    start_date: Optional[date] = Query(None, description="Report start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="Report end date (YYYY-MM-DD)"),
    location: Optional[str] = Query(None, description="Filter by inventory location"),
    category: Optional[str] = Query(None, description="Filter by supply category"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Dict:
    """
    Return a consumption report showing supply usage aggregated by category.

    The report is built from ``supply_requirements`` records, which represent
    forecasted demand (i.e., how much of each supply is expected to be consumed).

    Filters
    -------
    start_date : beginning of the reporting period (default: 30 days ago)
    end_date   : end of the reporting period (default: today)
    location   : filter inventory by warehouse / location field
    category   : filter supplies by category
    """
    start, end = _default_date_range(start_date, end_date)

    logger.info(
        "Consumption report requested by user=%s period=%s to %s location=%s category=%s",
        current_user.username, start, end, location, category,
    )

    # Aggregate required quantities grouped by category and supply
    query = (
        db.query(
            MedicalSupply.category,
            MedicalSupply.name.label("supply_name"),
            MedicalSupply.unit,
            func.sum(SupplyRequirement.required_quantity).label("total_required"),
            func.count(func.distinct(SupplyRequirement.requirement_date)).label("active_days"),
        )
        .join(SupplyRequirement, SupplyRequirement.supply_id == MedicalSupply.id)
        .filter(
            SupplyRequirement.requirement_date >= start,
            SupplyRequirement.requirement_date <= end,
        )
    )

    if category:
        query = query.filter(MedicalSupply.category == category)

    # Location filter: join inventory if location is provided
    if location:
        query = query.join(
            Inventory,
            Inventory.supply_id == MedicalSupply.id,
        ).filter(Inventory.location == location)

    rows = (
        query
        .group_by(MedicalSupply.category, MedicalSupply.name, MedicalSupply.unit)
        .order_by(MedicalSupply.category, MedicalSupply.name)
        .all()
    )

    # Group into category buckets
    category_map: Dict[str, Dict] = {}
    for row in rows:
        cat = row.category
        if cat not in category_map:
            category_map[cat] = {
                "category": cat,
                "total_required": 0,
                "supplies": [],
            }
        item = {
            "supply_name": row.supply_name,
            "unit": row.unit,
            "total_required": int(row.total_required or 0),
            "active_days": int(row.active_days or 0),
            "avg_daily_consumption": round(
                (row.total_required or 0) / max(int(row.active_days or 1), 1), 2
            ),
        }
        category_map[cat]["supplies"].append(item)
        category_map[cat]["total_required"] += item["total_required"]

    categories_list = sorted(category_map.values(), key=lambda c: -c["total_required"])
    grand_total = sum(c["total_required"] for c in categories_list)

    return {
        "report_type": "consumption",
        "period": {"start_date": str(start), "end_date": str(end)},
        "filters": {"location": location, "category": category},
        "summary": {
            "total_required_across_all_categories": grand_total,
            "categories_count": len(categories_list),
        },
        "categories": categories_list,
        "generated_at": datetime.utcnow().isoformat(),
    }


# ── Forecast Accuracy Report ──────────────────────────────────────────────────

@router.get("/forecast-accuracy")
async def get_forecast_accuracy_report(
    start_date: Optional[date] = Query(None, description="Report start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="Report end date (YYYY-MM-DD)"),
    disease_type: Optional[str] = Query(
        None,
        description="Filter by disease type: dengue_fever, seasonal_flu, respiratory_disease",
    ),
    model_used: Optional[str] = Query(
        None,
        description="Filter by model: xgboost, lstm, prophet, ensemble",
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Dict:
    """
    Return forecast accuracy metrics over time.

    Shows MAE, RMSE, and MAPE trends for each model, broken down by disease
    type.  Useful for monitoring model performance degradation over time.

    Filters
    -------
    start_date   : start of the reporting period (default: 30 days ago)
    end_date     : end of the reporting period (default: today)
    disease_type : restrict to a specific disease
    model_used   : restrict to a specific model name
    """
    start, end = _default_date_range(start_date, end_date)

    logger.info(
        "Forecast accuracy report requested by user=%s period=%s to %s "
        "disease_type=%s model_used=%s",
        current_user.username, start, end, disease_type, model_used,
    )

    query = (
        db.query(
            DiseaseForecast.forecast_date,
            DiseaseForecast.disease_type,
            DiseaseForecast.model_used,
            DiseaseForecast.model_accuracy_mae,
            DiseaseForecast.model_accuracy_rmse,
            DiseaseForecast.model_accuracy_mape,
            DiseaseForecast.predicted_cases,
            DiseaseForecast.confidence_lower,
            DiseaseForecast.confidence_upper,
        )
        .filter(
            DiseaseForecast.forecast_date >= start,
            DiseaseForecast.forecast_date <= end,
        )
    )

    if disease_type:
        query = query.filter(DiseaseForecast.disease_type == disease_type)
    if model_used:
        query = query.filter(DiseaseForecast.model_used == model_used)

    rows = query.order_by(DiseaseForecast.forecast_date).all()

    # Aggregate per model
    model_stats: Dict[str, Dict] = {}
    time_series: List[Dict] = []

    for row in rows:
        model = row.model_used or "unknown"
        if model not in model_stats:
            model_stats[model] = {
                "model": model,
                "sample_count": 0,
                "mae_values": [],
                "rmse_values": [],
                "mape_values": [],
            }
        stats = model_stats[model]
        stats["sample_count"] += 1
        if row.model_accuracy_mae is not None:
            stats["mae_values"].append(float(row.model_accuracy_mae))
        if row.model_accuracy_rmse is not None:
            stats["rmse_values"].append(float(row.model_accuracy_rmse))
        if row.model_accuracy_mape is not None:
            stats["mape_values"].append(float(row.model_accuracy_mape))

        time_series.append({
            "date": str(row.forecast_date),
            "disease_type": row.disease_type,
            "model": model,
            "predicted_cases": row.predicted_cases,
            "confidence_lower": row.confidence_lower,
            "confidence_upper": row.confidence_upper,
            "mae": float(row.model_accuracy_mae) if row.model_accuracy_mae is not None else None,
            "rmse": float(row.model_accuracy_rmse) if row.model_accuracy_rmse is not None else None,
            "mape": float(row.model_accuracy_mape) if row.model_accuracy_mape is not None else None,
        })

    # Compute averages per model
    model_summary = []
    for model, stats in model_stats.items():
        mae_vals = stats["mae_values"]
        rmse_vals = stats["rmse_values"]
        mape_vals = stats["mape_values"]
        model_summary.append({
            "model": model,
            "sample_count": stats["sample_count"],
            "avg_mae": round(sum(mae_vals) / len(mae_vals), 4) if mae_vals else None,
            "avg_rmse": round(sum(rmse_vals) / len(rmse_vals), 4) if rmse_vals else None,
            "avg_mape": round(sum(mape_vals) / len(mape_vals), 4) if mape_vals else None,
            "min_mae": round(min(mae_vals), 4) if mae_vals else None,
            "min_rmse": round(min(rmse_vals), 4) if rmse_vals else None,
        })

    # Best model by lowest avg MAPE
    best_model = None
    if model_summary:
        ranked = sorted(
            [m for m in model_summary if m["avg_mape"] is not None],
            key=lambda m: m["avg_mape"],
        )
        if ranked:
            best_model = ranked[0]["model"]

    return {
        "report_type": "forecast-accuracy",
        "period": {"start_date": str(start), "end_date": str(end)},
        "filters": {"disease_type": disease_type, "model_used": model_used},
        "summary": {
            "total_forecasts": len(rows),
            "models_evaluated": len(model_stats),
            "best_model_by_mape": best_model,
        },
        "model_performance": model_summary,
        "time_series": time_series,
        "generated_at": datetime.utcnow().isoformat(),
    }


# ── Inventory Turnover Report ─────────────────────────────────────────────────

@router.get("/inventory-turnover")
async def get_inventory_turnover_report(
    start_date: Optional[date] = Query(None, description="Report start date (YYYY-MM-DD)"),
    end_date: Optional[date] = Query(None, description="Report end date (YYYY-MM-DD)"),
    location: Optional[str] = Query(None, description="Filter by inventory location"),
    category: Optional[str] = Query(None, description="Filter by supply category"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Dict:
    """
    Return inventory turnover rates for all (or filtered) supplies.

    Turnover rate is calculated as:
        turnover_rate = total_required / avg_current_stock

    A higher value means the supply is consumed faster relative to on-hand
    stock.  Supplies with zero stock are marked as "out_of_stock".

    Filters
    -------
    start_date : start of the demand period used for the numerator (default: 30 days ago)
    end_date   : end of the demand period (default: today)
    location   : filter by inventory location
    category   : filter by supply category
    """
    start, end = _default_date_range(start_date, end_date)
    period_days = max((end - start).days, 1)

    logger.info(
        "Inventory turnover report requested by user=%s period=%s to %s "
        "location=%s category=%s",
        current_user.username, start, end, location, category,
    )

    # Build base query: join inventory → supply, optionally filter location
    inv_query = (
        db.query(
            Inventory.supply_id,
            Inventory.current_stock,
            Inventory.safety_stock,
            Inventory.location,
            MedicalSupply.name.label("supply_name"),
            MedicalSupply.category,
            MedicalSupply.unit,
            MedicalSupply.unit_price,
        )
        .join(MedicalSupply, MedicalSupply.id == Inventory.supply_id)
    )

    if location:
        inv_query = inv_query.filter(Inventory.location == location)
    if category:
        inv_query = inv_query.filter(MedicalSupply.category == category)

    inventory_rows = inv_query.all()

    # Get total demand per supply over the period
    demand_query = (
        db.query(
            SupplyRequirement.supply_id,
            func.sum(SupplyRequirement.required_quantity).label("total_required"),
        )
        .filter(
            SupplyRequirement.requirement_date >= start,
            SupplyRequirement.requirement_date <= end,
        )
        .group_by(SupplyRequirement.supply_id)
    )
    demand_map: Dict[int, int] = {
        row.supply_id: int(row.total_required or 0) for row in demand_query.all()
    }

    items = []
    for row in inventory_rows:
        sid = row.supply_id
        total_required = demand_map.get(sid, 0)
        current_stock = row.current_stock or 0
        safety_stock = row.safety_stock or 0
        unit_price = float(row.unit_price) if row.unit_price else 0.0

        if current_stock > 0:
            turnover_rate = round(total_required / current_stock, 4)
        else:
            turnover_rate = None  # out of stock

        days_of_supply = None
        if total_required > 0 and current_stock > 0:
            daily_demand = total_required / period_days
            days_of_supply = round(current_stock / daily_demand, 1)

        stock_status = "out_of_stock" if current_stock <= 0 else (
            "critical" if current_stock < safety_stock else "safe"
        )

        items.append({
            "supply_id": sid,
            "supply_name": row.supply_name,
            "category": row.category,
            "unit": row.unit,
            "location": row.location,
            "current_stock": current_stock,
            "safety_stock": safety_stock,
            "total_required_in_period": total_required,
            "turnover_rate": turnover_rate,
            "days_of_supply": days_of_supply,
            "stock_value": round(current_stock * unit_price, 2),
            "stock_status": stock_status,
        })

    # Sort: highest turnover first (None at end)
    items.sort(
        key=lambda x: (x["turnover_rate"] is None, -(x["turnover_rate"] or 0))
    )

    high_turnover = [i for i in items if i["turnover_rate"] is not None and i["turnover_rate"] > 1.0]
    out_of_stock = [i for i in items if i["stock_status"] == "out_of_stock"]
    avg_turnover = (
        round(sum(i["turnover_rate"] for i in items if i["turnover_rate"] is not None)
              / max(len([i for i in items if i["turnover_rate"] is not None]), 1), 4)
        if items else 0.0
    )

    return {
        "report_type": "inventory-turnover",
        "period": {"start_date": str(start), "end_date": str(end), "period_days": period_days},
        "filters": {"location": location, "category": category},
        "summary": {
            "total_items": len(items),
            "avg_turnover_rate": avg_turnover,
            "high_turnover_items": len(high_turnover),
            "out_of_stock_items": len(out_of_stock),
        },
        "items": items,
        "generated_at": datetime.utcnow().isoformat(),
    }


# ── Export Report ─────────────────────────────────────────────────────────────

class ReportExportRequest:
    """Simple holder – we use a Pydantic model below."""


from pydantic import BaseModel


class ExportReportRequest(BaseModel):
    """Request body for PDF/Excel report export."""

    report_type: str
    """One of: consumption, forecast-accuracy, inventory-turnover, dashboard-summary,
    epidemic, forecast, inventory, shortage, procurement."""

    format: str = "pdf"
    """Export format: 'pdf' or 'excel' (default 'pdf')."""

    start_date: Optional[date] = None
    end_date: Optional[date] = None
    location: Optional[str] = None
    category: Optional[str] = None
    disease_type: Optional[str] = None
    model_used: Optional[str] = None


@router.post("/export")
async def export_report(
    payload: ExportReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Response:
    """
    Generate a formatted PDF or Excel report for the requested report type.

    Supported report types (theo Module 8 - Smart Medical spec):
    - ``epidemic``            – Tình hình dịch bệnh
    - ``forecast``            – Dự báo ca bệnh
    - ``inventory``           – Tồn kho vật tư
    - ``shortage``            – Thiếu hụt vật tư
    - ``procurement``         – Đề xuất nhập kho
    - ``forecast-accuracy``   – Độ chính xác dự báo (alias 'accuracy')

    Legacy types vẫn được hỗ trợ:
    - ``consumption``, ``inventory-turnover``, ``dashboard-summary``

    `format` = 'pdf' (default) hoặc 'excel'.
    """
    SUPPORTED_TYPES = {
        # Smart Medical Module 8.2 — 6 loại chính
        "epidemic",
        "forecast",
        "inventory",
        "shortage",
        "procurement",
        "forecast-accuracy",
        # Legacy types
        "consumption",
        "inventory-turnover",
        "dashboard-summary",
    }
    if payload.report_type not in SUPPORTED_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported report_type '{payload.report_type}'. "
                   f"Must be one of: {sorted(SUPPORTED_TYPES)}",
        )

    if payload.format not in ("pdf", "excel"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported format '{payload.format}'. Use 'pdf' or 'excel'.",
        )

    logger.info(
        "Export report type=%s format=%s requested by user=%s",
        payload.report_type, payload.format, current_user.username,
    )

    start, end = _default_date_range(payload.start_date, payload.end_date)

    # Dispatch theo loại
    if payload.report_type in ("consumption",):
        data = await _build_consumption_data(db, start, end, payload.location, payload.category)
        return (
            _render_consumption_pdf(data, start, end)
            if payload.format == "pdf"
            else _render_consumption_excel(data, start, end)
        )

    if payload.report_type in ("forecast-accuracy",):
        data = await _build_accuracy_data(db, start, end, payload.disease_type, payload.model_used)
        return (
            _render_accuracy_pdf(data, start, end)
            if payload.format == "pdf"
            else _render_accuracy_excel(data, start, end)
        )

    if payload.report_type == "dashboard-summary":
        data = await _build_dashboard_summary_data(db)
        return (
            _render_dashboard_summary_pdf(data)
            if payload.format == "pdf"
            else _render_dashboard_summary_excel(data)
        )

    if payload.report_type == "inventory-turnover":
        data = await _build_turnover_data(db, start, end, payload.location, payload.category)
        return (
            _render_turnover_pdf(data, start, end)
            if payload.format == "pdf"
            else _render_turnover_excel(data, start, end)
        )

    # Smart Medical Module 8 — 5 loại mới
    if payload.report_type == "epidemic":
        data = await _build_epidemic_data(db, start, end, payload.disease_type, payload.location)
        return (
            _render_epidemic_pdf(data, start, end)
            if payload.format == "pdf"
            else _render_epidemic_excel(data, start, end)
        )

    if payload.report_type == "forecast":
        data = await _build_forecast_data(db, start, end, payload.disease_type, payload.location)
        return (
            _render_forecast_pdf(data, start, end)
            if payload.format == "pdf"
            else _render_forecast_excel(data, start, end)
        )

    if payload.report_type == "inventory":
        data = await _build_inventory_data(db, payload.category)
        return (
            _render_inventory_pdf(data)
            if payload.format == "pdf"
            else _render_inventory_excel(data)
        )

    if payload.report_type == "shortage":
        data = await _build_shortage_data(db, start, end, payload.disease_type)
        return (
            _render_shortage_pdf(data, start, end)
            if payload.format == "pdf"
            else _render_shortage_excel(data, start, end)
        )

    # procurement
    data = await _build_procurement_data(db, start, end, payload.disease_type)
    return (
        _render_procurement_pdf(data, start, end)
        if payload.format == "pdf"
        else _render_procurement_excel(data, start, end)
    )


# ── Internal data-fetching helpers (reused by export) ────────────────────────

async def _build_consumption_data(
    db: Session,
    start: date,
    end: date,
    location: Optional[str],
    category: Optional[str],
) -> List[Dict]:
    query = (
        db.query(
            MedicalSupply.category,
            MedicalSupply.name.label("supply_name"),
            MedicalSupply.unit,
            func.sum(SupplyRequirement.required_quantity).label("total_required"),
        )
        .join(SupplyRequirement, SupplyRequirement.supply_id == MedicalSupply.id)
        .filter(
            SupplyRequirement.requirement_date >= start,
            SupplyRequirement.requirement_date <= end,
        )
    )
    if category:
        query = query.filter(MedicalSupply.category == category)
    if location:
        query = query.join(Inventory, Inventory.supply_id == MedicalSupply.id).filter(
            Inventory.location == location
        )
    return query.group_by(
        MedicalSupply.category, MedicalSupply.name, MedicalSupply.unit
    ).order_by(MedicalSupply.category, MedicalSupply.name).all()


async def _build_accuracy_data(
    db: Session,
    start: date,
    end: date,
    disease_type: Optional[str],
    model_used: Optional[str],
) -> list:
    query = db.query(DiseaseForecast).filter(
        DiseaseForecast.forecast_date >= start,
        DiseaseForecast.forecast_date <= end,
    )
    if disease_type:
        query = query.filter(DiseaseForecast.disease_type == disease_type)
    if model_used:
        query = query.filter(DiseaseForecast.model_used == model_used)
    return query.order_by(DiseaseForecast.forecast_date).all()


async def _build_turnover_data(
    db: Session,
    start: date,
    end: date,
    location: Optional[str],
    category: Optional[str],
) -> list:
    inv_query = (
        db.query(
            Inventory.supply_id,
            Inventory.current_stock,
            Inventory.safety_stock,
            Inventory.location,
            MedicalSupply.name.label("supply_name"),
            MedicalSupply.category,
            MedicalSupply.unit,
        )
        .join(MedicalSupply, MedicalSupply.id == Inventory.supply_id)
    )
    if location:
        inv_query = inv_query.filter(Inventory.location == location)
    if category:
        inv_query = inv_query.filter(MedicalSupply.category == category)

    rows = inv_query.all()

    demand_map: Dict[int, int] = {
        row.supply_id: int(row.total_required or 0)
        for row in db.query(
            SupplyRequirement.supply_id,
            func.sum(SupplyRequirement.required_quantity).label("total_required"),
        )
        .filter(
            SupplyRequirement.requirement_date >= start,
            SupplyRequirement.requirement_date <= end,
        )
        .group_by(SupplyRequirement.supply_id)
        .all()
    }

    period_days = max((end - start).days, 1)
    result = []
    for row in rows:
        total_req = demand_map.get(row.supply_id, 0)
        cs = row.current_stock or 0
        turnover = round(total_req / cs, 4) if cs > 0 else None
        result.append({
            "supply_name": row.supply_name,
            "category": row.category,
            "unit": row.unit,
            "location": row.location or "",
            "current_stock": cs,
            "safety_stock": row.safety_stock or 0,
            "total_required": total_req,
            "turnover_rate": turnover,
        })
    result.sort(key=lambda x: (x["turnover_rate"] is None, -(x["turnover_rate"] or 0)))
    return result


# ── PDF rendering helpers ─────────────────────────────────────────────────────

# Đăng ký font Unicode (DejaVu Sans) một lần khi module load — hỗ trợ tiếng Việt.
# Nếu register thất bại, các renderer sẽ tự fallback về Helvetica (không có dấu).
PDF_FONT_REGULAR = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"
_FONT_REGISTERED = False


def _register_unicode_fonts() -> None:
    """Đăng ký DejaVu Sans (Regular + Bold) nếu chưa đăng ký.

    Các font này đi kèm matplotlib (là dependency hiện có) nên không cần
    phụ thuộc vào font hệ thống. Nếu không tìm thấy, giữ font Helvetica
    mặc định và log cảnh báo.
    """
    global _FONT_REGISTERED, PDF_FONT_REGULAR, PDF_FONT_BOLD
    if _FONT_REGISTERED:
        return
    try:
        import os
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        # Tìm thư mục font DejaVu trong matplotlib
        try:
            import matplotlib  # type: ignore

            font_dir = os.path.join(
                os.path.dirname(matplotlib.__file__),
                "mpl-data",
                "fonts",
                "ttf",
            )
            regular = os.path.join(font_dir, "DejaVuSans.ttf")
            bold = os.path.join(font_dir, "DejaVuSans-Bold.ttf")
        except Exception:
            regular = bold = ""

        # Fallback: font hệ thống macOS / Linux
        if not (os.path.exists(regular) and os.path.exists(bold)):
            for candidate in (
                "/Library/Fonts/Arial Unicode.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            ):
                if os.path.exists(candidate):
                    regular = candidate
                    bold = candidate  # cùng file dùng cho cả bold (sẽ bold giả lập)
                    break

        if not (regular and os.path.exists(regular)):
            logger.warning(
                "PDF font: không tìm thấy font Unicode, giữ Helvetica (không hỗ trợ tiếng Việt)."
            )
            _FONT_REGISTERED = True
            return

        pdfmetrics.registerFont(TTFont("DejaVuSans", regular))
        if bold and os.path.exists(bold) and bold != regular:
            pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold))
            PDF_FONT_BOLD = "DejaVuSans-Bold"
        else:
            PDF_FONT_BOLD = "DejaVuSans"
        PDF_FONT_REGULAR = "DejaVuSans"

        # Map family để các style có thể dùng <b>...</b> trong Paragraph
        from reportlab.pdfbase.pdfmetrics import registerFontFamily

        registerFontFamily(
            "DejaVuSans",
            normal="DejaVuSans",
            bold=PDF_FONT_BOLD,
            italic="DejaVuSans",
            boldItalic=PDF_FONT_BOLD,
        )
        logger.info("PDF font: registered DejaVu Sans Unicode fonts.")
    except Exception as exc:
        logger.warning("PDF font registration failed: %s", exc)
    finally:
        _FONT_REGISTERED = True


def _get_reportlab():
    """Import reportlab or raise a clean 500 error."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
        # Đăng ký font Unicode (idempotent)
        _register_unicode_fonts()
        return colors, A4, landscape, getSampleStyleSheet, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="reportlab library not available for PDF export",
        )


def _patch_styles_for_unicode(styles) -> None:
    """Áp dụng font Unicode lên tất cả các paragraph styles (Title, Heading, Normal, Italic...)."""
    for name in (
        "Normal",
        "BodyText",
        "Italic",
        "Title",
        "Heading1",
        "Heading2",
        "Heading3",
        "Heading4",
    ):
        try:
            style = styles[name]
            # Dùng bold hay regular tuỳ tên style
            if "Heading" in name or name == "Title":
                style.fontName = PDF_FONT_BOLD
            else:
                style.fontName = PDF_FONT_REGULAR
        except KeyError:
            pass


def _base_table_style(colors):
    """Return a shared base TableStyle list."""
    return [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 1), (-1, -1), PDF_FONT_REGULAR),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]


def _render_consumption_pdf(rows: list, start: date, end: date) -> Response:
    colors, A4, landscape, getSampleStyleSheet, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer = _get_reportlab()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    _patch_styles_for_unicode(styles)
    story = [
        Paragraph(
            f"Consumption Report – {start} to {end}  "
            f"(Generated {datetime.now().strftime('%Y-%m-%d %H:%M')})",
            styles["Title"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    table_data = [["Category", "Supply Name", "Unit", "Total Required"]]
    for row in rows:
        table_data.append([
            row.category,
            row.supply_name,
            row.unit,
            str(int(row.total_required or 0)),
        ])

    if len(table_data) == 1:
        story.append(Paragraph("No consumption data found for the selected period.", styles["Normal"]))
    else:
        col_widths = [5 * cm, 8 * cm, 3 * cm, 4 * cm]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle(_base_table_style(colors)))
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    filename = f"consumption_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _render_accuracy_pdf(forecasts: list, start: date, end: date) -> Response:
    colors, A4, landscape, getSampleStyleSheet, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer = _get_reportlab()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    _patch_styles_for_unicode(styles)
    story = [
        Paragraph(
            f"Forecast Accuracy Report – {start} to {end}  "
            f"(Generated {datetime.now().strftime('%Y-%m-%d %H:%M')})",
            styles["Title"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    table_data = [["Date", "Disease Type", "Model", "Predicted Cases", "MAE", "RMSE", "MAPE (%)"]]
    for fc in forecasts:
        table_data.append([
            str(fc.forecast_date),
            fc.disease_type or "",
            fc.model_used or "",
            str(fc.predicted_cases),
            f"{float(fc.model_accuracy_mae):.2f}" if fc.model_accuracy_mae is not None else "-",
            f"{float(fc.model_accuracy_rmse):.2f}" if fc.model_accuracy_rmse is not None else "-",
            f"{float(fc.model_accuracy_mape):.2f}" if fc.model_accuracy_mape is not None else "-",
        ])

    if len(table_data) == 1:
        story.append(Paragraph("No forecast accuracy data found for the selected period.", styles["Normal"]))
    else:
        col_widths = [3 * cm, 4.5 * cm, 3.5 * cm, 4 * cm, 3 * cm, 3 * cm, 3.5 * cm]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle(_base_table_style(colors)))
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    filename = f"forecast_accuracy_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _render_turnover_pdf(items: list, start: date, end: date) -> Response:
    colors, A4, landscape, getSampleStyleSheet, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer = _get_reportlab()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    _patch_styles_for_unicode(styles)
    story = [
        Paragraph(
            f"Inventory Turnover Report – {start} to {end}  "
            f"(Generated {datetime.now().strftime('%Y-%m-%d %H:%M')})",
            styles["Title"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    table_data = [["Supply Name", "Category", "Unit", "Location", "Current Stock", "Safety Stock", "Required", "Turnover Rate"]]
    for item in items:
        turnover = f"{item['turnover_rate']:.4f}" if item["turnover_rate"] is not None else "N/A"
        table_data.append([
            item["supply_name"],
            item["category"],
            item["unit"],
            item["location"],
            str(item["current_stock"]),
            str(item["safety_stock"]),
            str(item["total_required"]),
            turnover,
        ])

    if len(table_data) == 1:
        story.append(Paragraph("No inventory data found for the selected filters.", styles["Normal"]))
    else:
        col_widths = [5 * cm, 3.5 * cm, 2 * cm, 3 * cm, 3 * cm, 3 * cm, 3 * cm, 3.5 * cm]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        style = _base_table_style(colors)
        # Colour turnover column: green > 1, orange 0.5-1, red < 0.5
        for row_num, item in enumerate(items, 1):
            rate = item["turnover_rate"]
            if rate is not None:
                if rate >= 1.0:
                    c = colors.green
                elif rate >= 0.5:
                    c = colors.orange
                else:
                    c = colors.red
                style.append(("TEXTCOLOR", (7, row_num), (7, row_num), c))
                style.append(("FONTNAME", (7, row_num), (7, row_num), PDF_FONT_BOLD))
        tbl.setStyle(TableStyle(style))
        story.append(tbl)

    doc.build(story)
    buf.seek(0)
    filename = f"inventory_turnover_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Dashboard Summary Export ─────────────────────────────────────────────────


async def _build_dashboard_summary_data(db: Session) -> Dict:
    """Tổng hợp toàn bộ chỉ số đang hiển thị trên Dashboard."""
    from app.models.alert import Alert
    from app.models.disease_case import DiseaseCase

    today = date.today()
    first_of_this_month = today.replace(day=1)
    if first_of_this_month.month == 1:
        first_of_last_month = first_of_this_month.replace(
            year=first_of_this_month.year - 1, month=12,
        )
    else:
        first_of_last_month = first_of_this_month.replace(
            month=first_of_this_month.month - 1,
        )

    # KPI: tổng ca tháng này / tháng trước
    total_current = (
        db.query(func.coalesce(func.sum(DiseaseCase.case_count), 0))
        .filter(DiseaseCase.recorded_at >= first_of_this_month)
        .scalar()
        or 0
    )
    total_last = (
        db.query(func.coalesce(func.sum(DiseaseCase.case_count), 0))
        .filter(
            DiseaseCase.recorded_at >= first_of_last_month,
            DiseaseCase.recorded_at < first_of_this_month,
        )
        .scalar()
        or 0
    )
    cases_trend = (
        round(100.0 * (int(total_current) - int(total_last)) / int(total_last), 1)
        if total_last > 0 else 0.0
    )

    # KPI: dự báo tháng tới
    if first_of_this_month.month == 12:
        first_of_next = first_of_this_month.replace(
            year=first_of_this_month.year + 1, month=1,
        )
    else:
        first_of_next = first_of_this_month.replace(
            month=first_of_this_month.month + 1,
        )
    if first_of_next.month == 12:
        end_of_next = first_of_next.replace(
            year=first_of_next.year + 1, month=1,
        ) - timedelta(days=1)
    else:
        end_of_next = first_of_next.replace(
            month=first_of_next.month + 1,
        ) - timedelta(days=1)

    predicted_next = (
        db.query(func.coalesce(func.sum(DiseaseForecast.predicted_cases), 0))
        .filter(
            DiseaseForecast.forecast_date >= first_of_next,
            DiseaseForecast.forecast_date <= end_of_next,
        )
        .scalar()
        or 0
    )
    predicted_trend = (
        round(100.0 * (int(predicted_next) - int(total_current)) / int(total_current), 1)
        if total_current > 0 else 0.0
    )

    # KPI: số vật tư thiếu hụt
    shortage_count = (
        db.query(func.count(Alert.id))
        .filter(Alert.is_resolved == False)  # noqa: E712
        .scalar()
        or 0
    )

    # Mức nguy cơ chung (cùng công thức như endpoint /dashboard/summary)
    if predicted_trend >= 15 and shortage_count >= 5:
        overall_risk = "Cao"
    elif predicted_trend >= 5 or shortage_count >= 2:
        overall_risk = "Trung bình"
    else:
        overall_risk = "Thấp"

    # Xu hướng 6 tháng (this year + last year)
    trend_rows = []
    for i in range(5, -1, -1):
        year = today.year
        month = today.month - i
        while month <= 0:
            month += 12
            year -= 1
        start_m = date(year, month, 1)
        end_m = date(year + (1 if month == 12 else 0), (month % 12) + 1, 1) - timedelta(days=1)
        this_y = (
            db.query(func.coalesce(func.sum(DiseaseCase.case_count), 0))
            .filter(DiseaseCase.recorded_at >= start_m, DiseaseCase.recorded_at <= end_m)
            .scalar()
            or 0
        )
        last_y = (
            db.query(func.coalesce(func.sum(DiseaseCase.case_count), 0))
            .filter(
                DiseaseCase.recorded_at >= start_m.replace(year=start_m.year - 1),
                DiseaseCase.recorded_at <= end_m.replace(year=end_m.year - 1),
            )
            .scalar()
            or 0
        )
        trend_rows.append({"month": f"T{month}", "this_year": int(this_y), "last_year": int(last_y)})

    # Top 5 vật tư demand vs stock
    end_demand = today + timedelta(days=60)
    demand_rows_raw = (
        db.query(
            MedicalSupply.id,
            MedicalSupply.name,
            MedicalSupply.unit,
            func.coalesce(func.sum(SupplyRequirement.required_quantity), 0).label("demand"),
        )
        .join(SupplyRequirement, SupplyRequirement.supply_id == MedicalSupply.id)
        .filter(
            SupplyRequirement.requirement_date >= today,
            SupplyRequirement.requirement_date <= end_demand,
        )
        .group_by(MedicalSupply.id, MedicalSupply.name, MedicalSupply.unit)
        .order_by(func.sum(SupplyRequirement.required_quantity).desc())
        .limit(5)
        .all()
    )
    demand_rows = []
    for row in demand_rows_raw:
        stock = (
            db.query(func.coalesce(func.sum(Inventory.current_stock), 0))
            .filter(Inventory.supply_id == row.id)
            .scalar()
            or 0
        )
        demand_rows.append({
            "supply_name": row.name,
            "unit": row.unit,
            "demand": int(row.demand),
            "stock": int(stock),
        })

    # Bảng cảnh báo top 5 (critical + high)
    alert_rows = (
        db.query(Alert)
        .filter(
            Alert.is_resolved == False,  # noqa: E712
            Alert.severity.in_(["critical", "high"]),
        )
        .order_by(Alert.severity.asc(), Alert.created_at.desc())
        .limit(5)
        .all()
    )
    alerts_list = [
        {
            "supply_name": a.supply.name if a.supply else f"Supply #{a.supply_id}",
            "current_stock": a.current_stock or 0,
            "required_stock": a.required_stock or 0,
            "severity": a.severity,
        }
        for a in alert_rows
    ]

    return {
        "as_of": today.isoformat(),
        "month_label": first_of_this_month.strftime("%m/%Y"),
        "kpi": {
            "total_cases_current": int(total_current),
            "cases_trend_pct": cases_trend,
            "predicted_cases_next_month": int(predicted_next),
            "predicted_trend_pct": predicted_trend,
            "shortage_supplies_count": int(shortage_count),
            "overall_risk": overall_risk,
        },
        "case_trend": trend_rows,
        "demand_vs_stock": demand_rows,
        "alerts": alerts_list,
    }


def _render_dashboard_summary_pdf(data: Dict) -> Response:
    """Render báo cáo tổng quan Dashboard ra file PDF."""
    colors, A4, landscape, getSampleStyleSheet, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer = _get_reportlab()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    _patch_styles_for_unicode(styles)
    story = []

    # Title
    story.append(Paragraph(
        f"Báo cáo Dashboard tổng quan - tháng {data['month_label']}",
        styles["Title"],
    ))
    story.append(Paragraph(
        f"Thời điểm xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
        f"Dữ liệu đến ngày: {data['as_of']}",
        styles["Italic"],
    ))
    story.append(Spacer(1, 0.5 * cm))

    base_style = _base_table_style(colors)

    # 1. KPI table
    story.append(Paragraph("<b>I. Chỉ số tổng quan (KPI)</b>", styles["Heading2"]))
    kpi = data["kpi"]
    kpi_table = Table(
        [
            ["Chỉ số", "Giá trị", "Xu hướng"],
            [
                "Tổng số ca hiện tại",
                f"{kpi['total_cases_current']:,}",
                f"{kpi['cases_trend_pct']:+.1f}%",
            ],
            [
                "Số ca dự báo tháng tới",
                f"{kpi['predicted_cases_next_month']:,}",
                f"{kpi['predicted_trend_pct']:+.1f}%",
            ],
            [
                "Vật tư thiếu hụt",
                f"{kpi['shortage_supplies_count']:,} mục",
                "—",
            ],
            ["Mức nguy cơ chung", kpi["overall_risk"], "—"],
        ],
        colWidths=[8 * cm, 5 * cm, 4 * cm],
        repeatRows=1,
    )
    kpi_table.setStyle(TableStyle(base_style))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5 * cm))

    # 2. Case trend table
    story.append(Paragraph("<b>II. Xu hướng ca bệnh 6 tháng</b>", styles["Heading2"]))
    trend_data = [["Tháng", "Năm nay", "Năm trước"]]
    for r in data["case_trend"]:
        trend_data.append([r["month"], f"{r['this_year']:,}", f"{r['last_year']:,}"])
    trend_table = Table(trend_data, colWidths=[4 * cm, 5 * cm, 5 * cm], repeatRows=1)
    trend_table.setStyle(TableStyle(base_style))
    story.append(trend_table)
    story.append(Spacer(1, 0.5 * cm))

    # 3. Demand vs Stock table
    story.append(Paragraph("<b>III. Nhu cầu vs Tồn kho (Top 5)</b>", styles["Heading2"]))

    # Style nhỏ cho text dài trong cell, tự wrap
    from reportlab.lib.styles import ParagraphStyle

    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName=PDF_FONT_REGULAR,
        fontSize=8,
        leading=10,
        wordWrap="CJK",  # cho phép wrap ở mọi vị trí, kể cả không có space
    )

    if data["demand_vs_stock"]:
        ds_data = [["Vật tư", "Đơn vị", "Tồn kho", "Nhu cầu"]]
        for r in data["demand_vs_stock"]:
            ds_data.append([
                Paragraph(r["supply_name"], cell_style),
                Paragraph(r["unit"] or "", cell_style),
                f"{r['stock']:,}",
                f"{r['demand']:,}",
            ])
        ds_table = Table(ds_data, colWidths=[8.5 * cm, 2 * cm, 2.7 * cm, 2.8 * cm], repeatRows=1)
        ds_table.setStyle(TableStyle(base_style))
        story.append(ds_table)
    else:
        story.append(Paragraph("Không có dữ liệu nhu cầu trong 60 ngày tới.", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    # 4. Alerts table
    story.append(Paragraph("<b>IV. Cảnh báo thiếu hụt vật tư (Top 5)</b>", styles["Heading2"]))
    if data["alerts"]:
        severity_label = {"critical": "Nguy hiểm", "high": "Cần nhập", "medium": "Cảnh báo"}
        a_data = [["Vật tư", "Tồn hiện tại", "Định mức", "Trạng thái"]]
        for a in data["alerts"]:
            a_data.append([
                Paragraph(a["supply_name"], cell_style),
                f"{a['current_stock']:,}",
                f"{a['required_stock']:,}",
                severity_label.get(a["severity"], a["severity"]),
            ])
        a_table = Table(a_data, colWidths=[8.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm], repeatRows=1)
        style = list(base_style)
        # Color severity column
        for row_num, a in enumerate(data["alerts"], 1):
            sev = a["severity"]
            if sev == "critical":
                style.append(("TEXTCOLOR", (3, row_num), (3, row_num), colors.red))
            elif sev == "high":
                style.append(("TEXTCOLOR", (3, row_num), (3, row_num), colors.orange))
            else:
                style.append(("TEXTCOLOR", (3, row_num), (3, row_num), colors.HexColor("#B45309")))
            style.append(("FONTNAME", (3, row_num), (3, row_num), PDF_FONT_BOLD))
        a_table.setStyle(TableStyle(style))
        story.append(a_table)
    else:
        story.append(Paragraph("Không có cảnh báo nào.", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    filename = f"dashboard_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



# ═══════════════════════════════════════════════════════════════════════════
# Module 8 — Smart Medical Reports: Data Builders cho 5 loại mới
# ═══════════════════════════════════════════════════════════════════════════


async def _build_epidemic_data(
    db: Session,
    start: date,
    end: date,
    disease_type: Optional[str],
    location: Optional[str],
) -> Dict:
    """Báo cáo Tình hình dịch bệnh: số ca theo tháng × bệnh × khu vực."""
    from app.models.disease_case import DiseaseCase

    q = db.query(
        func.strftime("%m/%Y", DiseaseCase.recorded_at).label("month"),
        DiseaseCase.disease_type,
        DiseaseCase.location,
        func.sum(DiseaseCase.case_count).label("total"),
    ).filter(
        DiseaseCase.recorded_at >= datetime(start.year, start.month, start.day),
        DiseaseCase.recorded_at <= datetime(end.year, end.month, end.day),
    )
    if disease_type:
        q = q.filter(DiseaseCase.disease_type == disease_type)
    if location:
        q = q.filter(DiseaseCase.location == location)
    rows = q.group_by(
        func.strftime("%m/%Y", DiseaseCase.recorded_at),
        DiseaseCase.disease_type,
        DiseaseCase.location,
    ).order_by(
        func.strftime("%Y-%m", DiseaseCase.recorded_at),
        DiseaseCase.disease_type,
    ).all()

    items = [
        {
            "month": r.month,
            "disease_type": r.disease_type,
            "disease_label": _vi_disease(r.disease_type),
            "location": r.location,
            "total": int(r.total or 0),
        }
        for r in rows
    ]
    total_cases = sum(it["total"] for it in items)
    return {
        "items": items,
        "total_cases": total_cases,
        "filters": {"disease_type": disease_type, "location": location},
    }


async def _build_forecast_data(
    db: Session,
    start: date,
    end: date,
    disease_type: Optional[str],
    location: Optional[str],
) -> Dict:
    """Báo cáo Dự báo ca bệnh: predicted_cases + risk + explanation."""
    from app.models.disease_forecast import DiseaseForecast

    q = db.query(DiseaseForecast).filter(
        DiseaseForecast.forecast_date >= start,
        DiseaseForecast.forecast_date <= end,
    )
    if disease_type:
        q = q.filter(DiseaseForecast.disease_type == disease_type)
    if location:
        q = q.filter(DiseaseForecast.location == location)
    rows = q.order_by(DiseaseForecast.forecast_date.desc()).all()

    risk_label = {
        "low": "Thấp",
        "medium": "Trung bình",
        "high": "Cao",
        "very_high": "Rất cao",
    }

    items = [
        {
            "month": r.forecast_date.strftime("%m/%Y") if r.forecast_date else "—",
            "disease_label": _vi_disease(r.disease_type),
            "location": r.location or "Toàn thành phố",
            "predicted_cases": r.predicted_cases or 0,
            "baseline_cases": r.baseline_cases or 0,
            "risk_level": r.risk_level or "",
            "risk_label": risk_label.get(r.risk_level or "", "—"),
            "explanation": (r.explanation or "")[:300],
        }
        for r in rows
    ]
    return {
        "items": items,
        "filters": {"disease_type": disease_type, "location": location},
    }


async def _build_inventory_data(
    db: Session,
    category: Optional[str],
) -> Dict:
    """Báo cáo Tồn kho: vật tư + tồn kho + ngưỡng AT + trạng thái."""
    from app.models.medical_supply import MedicalSupply
    from app.models.inventory import Inventory as InventoryModel

    q = db.query(
        Inventory.id,
        Inventory.current_stock,
        Inventory.safety_stock,
        Inventory.expiry_date,
        MedicalSupply.id.label("supply_id"),
        MedicalSupply.name,
        MedicalSupply.category,
        MedicalSupply.unit,
    ).join(MedicalSupply, MedicalSupply.id == Inventory.supply_id)
    if category:
        q = q.filter(MedicalSupply.category == category)
    rows = q.order_by(MedicalSupply.name).all()

    items = []
    for r in rows:
        cur = r.current_stock or 0
        saf = r.safety_stock or 0
        if cur <= 0 or (saf > 0 and cur < saf * 0.3):
            status_text = "Cần nhập gấp"
        elif cur <= saf:
            status_text = "Dưới ngưỡng"
        else:
            status_text = "Bình thường"
        items.append(
            {
                "supply_name": r.name,
                "category": _vi_category(r.category),
                "unit": r.unit,
                "current_stock": cur,
                "safety_stock": saf,
                "status": status_text,
                "expiry_date": r.expiry_date.strftime("%d/%m/%Y") if r.expiry_date else "",
            }
        )
    return {
        "items": items,
        "summary": {
            "total": len(items),
            "critical": sum(1 for x in items if x["status"] == "Cần nhập gấp"),
            "low": sum(1 for x in items if x["status"] == "Dưới ngưỡng"),
            "safe": sum(1 for x in items if x["status"] == "Bình thường"),
        },
    }


async def _build_shortage_data(
    db: Session,
    start: date,
    end: date,
    disease_type: Optional[str],
) -> Dict:
    """Báo cáo Thiếu hụt vật tư: vật tư có shortage > 0 trong kỳ."""
    from app.models.medical_supply import MedicalSupply
    from app.models.inventory import Inventory as InventoryModel

    demand_q = db.query(
        SupplyRequirement.supply_id,
        func.sum(SupplyRequirement.required_quantity).label("total_required"),
    ).filter(
        SupplyRequirement.requirement_date >= start,
        SupplyRequirement.requirement_date <= end,
    )
    if disease_type:
        demand_q = demand_q.filter(SupplyRequirement.disease_type == disease_type)
    demand_map = {row.supply_id: int(row.total_required or 0) for row in demand_q.group_by(SupplyRequirement.supply_id).all()}

    if not demand_map:
        return {"items": [], "summary": {"total": 0, "total_shortage": 0}}

    supplies = (
        db.query(MedicalSupply, Inventory.current_stock)
        .outerjoin(Inventory, Inventory.supply_id == MedicalSupply.id)
        .filter(MedicalSupply.id.in_(demand_map.keys()))
        .all()
    )

    items = []
    for s, stock in supplies:
        demand = demand_map.get(s.id, 0)
        cur = int(stock or 0)
        shortage = max(0, demand - cur)
        if shortage <= 0:
            continue
        items.append(
            {
                "supply_name": s.name,
                "category": _vi_category(s.category),
                "unit": s.unit,
                "demand": demand,
                "stock": cur,
                "shortage": shortage,
            }
        )
    items.sort(key=lambda x: -x["shortage"])
    return {
        "items": items,
        "summary": {
            "total": len(items),
            "total_shortage": sum(x["shortage"] for x in items),
        },
    }


async def _build_procurement_data(
    db: Session,
    start: date,
    end: date,
    disease_type: Optional[str],
) -> Dict:
    """Báo cáo Đề xuất nhập kho: vật tư cần nhập + lý do."""
    from app.models.medical_supply import MedicalSupply
    from app.models.inventory import Inventory as InventoryModel
    from app.models.system_config import SystemConfig

    # Lấy safety rate từ admin
    cfg = db.query(SystemConfig).filter(SystemConfig.config_key == "admin.safety_rate").first()
    try:
        safety_rate = float(cfg.config_value) if cfg else 0.15
    except (TypeError, ValueError):
        safety_rate = 0.15

    demand_q = db.query(
        SupplyRequirement.supply_id,
        func.sum(SupplyRequirement.required_quantity).label("total_required"),
    ).filter(
        SupplyRequirement.requirement_date >= start,
        SupplyRequirement.requirement_date <= end,
    )
    if disease_type:
        demand_q = demand_q.filter(SupplyRequirement.disease_type == disease_type)
    demand_map = {row.supply_id: int(row.total_required or 0) for row in demand_q.group_by(SupplyRequirement.supply_id).all()}

    if not demand_map:
        return {"items": [], "summary": {"total": 0, "total_order": 0}, "safety_rate": safety_rate}

    supplies = (
        db.query(MedicalSupply, Inventory.current_stock)
        .outerjoin(Inventory, Inventory.supply_id == MedicalSupply.id)
        .filter(MedicalSupply.id.in_(demand_map.keys()))
        .all()
    )

    items = []
    for s, stock in supplies:
        demand = demand_map.get(s.id, 0)
        cur = int(stock or 0)
        safety = round(demand * safety_rate)
        recommended = max(0, demand + safety - cur)
        if recommended <= 0:
            continue
        ratio = (cur / demand) if demand > 0 else 1
        if ratio < 0.1:
            reason = "Tồn < 10% nhu cầu — nguy hiểm"
        elif ratio < 0.25:
            reason = "Tồn 10–25% nhu cầu — cảnh báo"
        else:
            reason = "Bù dự phòng theo chính sách"
        items.append(
            {
                "supply_name": s.name,
                "category": _vi_category(s.category),
                "unit": s.unit,
                "demand": demand,
                "stock": cur,
                "recommended": recommended,
                "reason": reason,
            }
        )
    items.sort(key=lambda x: -x["recommended"])
    return {
        "items": items,
        "summary": {
            "total": len(items),
            "total_order": sum(x["recommended"] for x in items),
        },
        "safety_rate": safety_rate,
    }


def _ascii_filename(s: str) -> str:
    """Chuyển string tiếng Việt sang dạng ASCII không dấu để dùng trong HTTP header."""
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_only = "".join(c for c in nfkd if not unicodedata.combining(c))
    # Bỏ ký tự đặc biệt không phải ASCII còn lại
    return ascii_only.encode("ascii", errors="ignore").decode("ascii") or "report"


def _vi_disease(key: str) -> str:
    return {
        "dengue_fever": "Sốt xuất huyết",
        "seasonal_flu": "Cúm mùa",
        "respiratory_disease": "Bệnh hô hấp",
        "viral_infection": "Nhiễm virus",
    }.get(key or "", key or "—")


def _vi_category(key: str) -> str:
    return {
        "medicine": "Thuốc",
        "mask": "Khẩu trang",
        "glove": "Găng tay",
        "test_kit": "Kit XN",
        "disinfectant": "Hoá chất",
        "iv_fluid": "Dịch truyền",
        "other": "Khác",
    }.get(key or "", key or "—")



# ═══════════════════════════════════════════════════════════════════════════
# PDF Renderers cho 5 loại mới (epidemic / forecast / inventory / shortage / procurement)
# ═══════════════════════════════════════════════════════════════════════════


def _generic_pdf(
    title: str,
    period_label: str,
    headers: list[str],
    rows: list[list[str]],
    col_widths_cm: list[float],
    summary_lines: Optional[list[str]] = None,
    landscape_mode: bool = True,
) -> Response:
    """Helper: render bảng đơn giản ra PDF với font Unicode.

    col_widths_cm: list số float — đơn vị cm (sẽ được nhân với 1*cm bên trong).
    """
    colors, A4, _landscape, getSampleStyleSheet, cm, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer = _get_reportlab()

    pagesize = _landscape(A4) if landscape_mode else A4
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    _patch_styles_for_unicode(styles)
    story = [
        Paragraph(title, styles["Title"]),
        Paragraph(
            f"Kỳ báo cáo: {period_label} | Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles["Italic"],
        ),
        Spacer(1, 0.4 * cm),
    ]

    # Wrap text dài trong cell bằng Paragraph
    from reportlab.lib.styles import ParagraphStyle
    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontName=PDF_FONT_REGULAR,
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )

    if not rows:
        story.append(Paragraph("Không có dữ liệu trong kỳ báo cáo này.", styles["Normal"]))
    else:
        # Convert col_widths_cm sang ReportLab unit
        col_widths = [w * cm for w in col_widths_cm]
        # Convert rows: nếu cell quá dài thì wrap
        formatted_rows: list[list] = [headers]
        for r in rows:
            formatted_rows.append([
                Paragraph(str(c), cell_style) if isinstance(c, str) and len(c) > 30 else str(c)
                for c in r
            ])
        tbl = Table(formatted_rows, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle(_base_table_style(colors)))
        story.append(tbl)

    if summary_lines:
        story.append(Spacer(1, 0.4 * cm))
        for line in summary_lines:
            story.append(Paragraph(line, styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    safe = title.replace(" ", "_").replace("/", "-")[:40]
    # Encode tên file theo RFC 5987 vì HTTP headers chỉ chấp nhận Latin-1
    from urllib.parse import quote as _quote
    filename_ascii = _ascii_filename(safe) + f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_utf8 = _quote(safe + f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_ascii}"; '
                f"filename*=UTF-8''{filename_utf8}"
            )
        },
    )


def _render_epidemic_pdf(data: Dict, start: date, end: date) -> Response:
    rows = [
        [it["month"], it["disease_label"], it["location"], f"{it['total']:,}"]
        for it in data["items"]
    ]
    return _generic_pdf(
        title="Báo cáo Tình hình Dịch bệnh",
        period_label=f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}",
        headers=["Tháng", "Bệnh", "Khu vực", "Số ca"],
        rows=rows,
        col_widths_cm=[3, 5, 8, 4],
        summary_lines=[f"<b>Tổng số ca trong kỳ:</b> {data['total_cases']:,}"],
        landscape_mode=False,
    )


def _render_forecast_pdf(data: Dict, start: date, end: date) -> Response:
    rows = [
        [
            it["month"],
            it["disease_label"],
            it["location"],
            f"{it['predicted_cases']:,}",
            it["risk_label"],
            it["explanation"][:200],
        ]
        for it in data["items"]
    ]
    return _generic_pdf(
        title="Báo cáo Dự báo Ca bệnh",
        period_label=f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}",
        headers=["Tháng", "Bệnh", "Khu vực", "Dự báo", "Mức nguy cơ", "Lý do"],
        rows=rows,
        col_widths_cm=[2.5, 4, 4.5, 2.5, 2.5, 9],
    )


def _render_inventory_pdf(data: Dict) -> Response:
    rows = [
        [
            it["supply_name"],
            it["category"],
            it["unit"],
            f"{it['current_stock']:,}",
            f"{it['safety_stock']:,}",
            it["status"],
            it["expiry_date"],
        ]
        for it in data["items"]
    ]
    s = data["summary"]
    return _generic_pdf(
        title="Báo cáo Tồn kho Vật tư",
        period_label=datetime.now().strftime("%d/%m/%Y"),
        headers=["Vật tư", "Loại", "ĐVT", "Tồn kho", "Ngưỡng AT", "Trạng thái", "Hạn dùng"],
        rows=rows,
        col_widths_cm=[7, 3, 1.8, 2.2, 2.2, 3, 2.5],
        summary_lines=[
            f"<b>Tổng vật tư:</b> {s['total']:,} | "
            f"<b>An toàn:</b> {s['safe']:,} | "
            f"<b>Dưới ngưỡng:</b> {s['low']:,} | "
            f"<b>Cần nhập gấp:</b> {s['critical']:,}"
        ],
    )


def _render_shortage_pdf(data: Dict, start: date, end: date) -> Response:
    rows = [
        [
            it["supply_name"],
            it["category"],
            it["unit"],
            f"{it['demand']:,}",
            f"{it['stock']:,}",
            f"{it['shortage']:,}",
        ]
        for it in data["items"]
    ]
    s = data["summary"]
    return _generic_pdf(
        title="Báo cáo Thiếu hụt Vật tư",
        period_label=f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}",
        headers=["Vật tư", "Loại", "ĐVT", "Nhu cầu", "Tồn kho", "Mức thiếu"],
        rows=rows,
        col_widths_cm=[7.5, 3, 2, 2.5, 2.5, 2.5],
        summary_lines=[
            f"<b>Số vật tư thiếu:</b> {s['total']:,} | "
            f"<b>Tổng lượng thiếu:</b> {s['total_shortage']:,}"
        ],
    )


def _render_procurement_pdf(data: Dict, start: date, end: date) -> Response:
    rows = [
        [
            it["supply_name"],
            it["category"],
            it["unit"],
            f"{it['demand']:,}",
            f"{it['stock']:,}",
            f"{it['recommended']:,}",
            it["reason"],
        ]
        for it in data["items"]
    ]
    s = data["summary"]
    rate_pct = round(data.get("safety_rate", 0.15) * 100)
    return _generic_pdf(
        title="Báo cáo Đề xuất Nhập kho",
        period_label=f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}",
        headers=["Vật tư", "Loại", "ĐVT", "Nhu cầu", "Tồn kho", "SL nhập", "Lý do"],
        rows=rows,
        col_widths_cm=[5.5, 2.5, 1.8, 2.2, 2.2, 2.5, 6],
        summary_lines=[
            f"<b>Số vật tư đề xuất nhập:</b> {s['total']:,} | "
            f"<b>Tổng SL đề xuất:</b> {s['total_order']:,} | "
            f"<b>Hệ số dự phòng:</b> {rate_pct}%"
        ],
    )


# ═══════════════════════════════════════════════════════════════════════════
# Excel Renderers — dùng openpyxl
# ═══════════════════════════════════════════════════════════════════════════


def _generic_excel(
    sheet_title: str,
    headers: list[str],
    rows: list[list],
    column_widths: list[int],
    filename_prefix: str,
    title_line: Optional[str] = None,
) -> Response:
    """Helper: render bảng đơn giản ra Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl chưa được cài",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    start_row = 1
    if title_line:
        ws.cell(row=1, column=1, value=title_line)
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        start_row = 3

    for col_num, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col_num, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center

    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for col_num, w in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote as _quote
    name_base = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{_ascii_filename(name_base)}"; '
                f"filename*=UTF-8''{_quote(name_base)}"
            )
        },
    )


# ── Excel: 5 loại mới ──────────────────────────────────────────────────────


def _render_epidemic_excel(data: Dict, start: date, end: date) -> Response:
    rows = [
        [it["month"], it["disease_label"], it["location"], it["total"]]
        for it in data["items"]
    ]
    return _generic_excel(
        sheet_title="Tình hình dịch bệnh",
        headers=["Tháng", "Bệnh", "Khu vực", "Số ca"],
        rows=rows,
        column_widths=[12, 24, 28, 12],
        filename_prefix="bao_cao_dich_benh",
        title_line=f"Báo cáo Tình hình Dịch bệnh — {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}",
    )


def _render_forecast_excel(data: Dict, start: date, end: date) -> Response:
    rows = [
        [
            it["month"],
            it["disease_label"],
            it["location"],
            it["predicted_cases"],
            it["baseline_cases"],
            it["risk_label"],
            it["explanation"],
        ]
        for it in data["items"]
    ]
    return _generic_excel(
        sheet_title="Dự báo ca bệnh",
        headers=["Tháng", "Bệnh", "Khu vực", "Số ca dự báo", "Ca nền", "Mức nguy cơ", "Lý do dự báo"],
        rows=rows,
        column_widths=[10, 22, 24, 14, 12, 14, 60],
        filename_prefix="bao_cao_du_bao",
        title_line=f"Báo cáo Dự báo Ca bệnh — {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}",
    )


def _render_inventory_excel(data: Dict) -> Response:
    rows = [
        [
            it["supply_name"],
            it["category"],
            it["unit"],
            it["current_stock"],
            it["safety_stock"],
            it["status"],
            it["expiry_date"],
        ]
        for it in data["items"]
    ]
    return _generic_excel(
        sheet_title="Tồn kho",
        headers=["Vật tư", "Loại", "ĐVT", "Tồn kho", "Ngưỡng AT", "Trạng thái", "Hạn dùng"],
        rows=rows,
        column_widths=[40, 14, 8, 12, 12, 16, 14],
        filename_prefix="bao_cao_ton_kho",
        title_line=f"Báo cáo Tồn kho Vật tư — {datetime.now().strftime('%d/%m/%Y')}",
    )


def _render_shortage_excel(data: Dict, start: date, end: date) -> Response:
    rows = [
        [
            it["supply_name"],
            it["category"],
            it["unit"],
            it["demand"],
            it["stock"],
            it["shortage"],
        ]
        for it in data["items"]
    ]
    return _generic_excel(
        sheet_title="Thiếu hụt",
        headers=["Vật tư", "Loại", "ĐVT", "Nhu cầu", "Tồn kho", "Mức thiếu"],
        rows=rows,
        column_widths=[40, 14, 8, 14, 14, 14],
        filename_prefix="bao_cao_thieu_hut",
        title_line=f"Báo cáo Thiếu hụt Vật tư — {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}",
    )


def _render_procurement_excel(data: Dict, start: date, end: date) -> Response:
    rows = [
        [
            it["supply_name"],
            it["category"],
            it["unit"],
            it["demand"],
            it["stock"],
            it["recommended"],
            it["reason"],
        ]
        for it in data["items"]
    ]
    return _generic_excel(
        sheet_title="Đề xuất nhập kho",
        headers=["Vật tư", "Loại", "ĐVT", "Nhu cầu", "Tồn kho", "SL nhập", "Lý do"],
        rows=rows,
        column_widths=[36, 14, 8, 14, 14, 14, 36],
        filename_prefix="bao_cao_de_xuat",
        title_line=f"Báo cáo Đề xuất Nhập kho — {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}",
    )


# ── Excel cho 4 loại legacy ─────────────────────────────────────────────────


def _render_consumption_excel(rows_query, start: date, end: date) -> Response:
    """rows_query là kết quả SQL — chuyển sang list."""
    rows = [
        [r.category, r.supply_name, r.unit, int(r.total_required or 0)]
        for r in rows_query
    ]
    return _generic_excel(
        sheet_title="Tiêu thụ",
        headers=["Loại", "Vật tư", "ĐVT", "Tổng yêu cầu"],
        rows=rows,
        column_widths=[14, 40, 8, 14],
        filename_prefix="bao_cao_tieu_thu",
        title_line=f"Báo cáo Tiêu thụ — {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}",
    )


def _render_accuracy_excel(forecasts, start: date, end: date) -> Response:
    rows = [
        [
            str(fc.forecast_date),
            fc.disease_type or "",
            fc.model_used or "",
            fc.predicted_cases,
            float(fc.model_accuracy_mae) if fc.model_accuracy_mae is not None else None,
            float(fc.model_accuracy_rmse) if fc.model_accuracy_rmse is not None else None,
            float(fc.model_accuracy_mape) if fc.model_accuracy_mape is not None else None,
        ]
        for fc in forecasts
    ]
    return _generic_excel(
        sheet_title="Độ chính xác",
        headers=["Ngày", "Bệnh", "Mô hình", "Dự báo", "MAE", "RMSE", "MAPE %"],
        rows=rows,
        column_widths=[12, 22, 14, 12, 12, 12, 12],
        filename_prefix="bao_cao_chinh_xac",
        title_line=f"Báo cáo Độ chính xác Dự báo — {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}",
    )


def _render_turnover_excel(items: list, start: date, end: date) -> Response:
    rows = [
        [
            it["supply_name"],
            it["category"],
            it["unit"],
            it["location"],
            it["current_stock"],
            it["safety_stock"],
            it["total_required"],
            it["turnover_rate"],
        ]
        for it in items
    ]
    return _generic_excel(
        sheet_title="Vòng quay",
        headers=["Vật tư", "Loại", "ĐVT", "Khu vực", "Tồn kho", "Ngưỡng AT", "Yêu cầu", "Vòng quay"],
        rows=rows,
        column_widths=[36, 14, 8, 14, 12, 14, 12, 14],
        filename_prefix="bao_cao_vong_quay",
        title_line=f"Báo cáo Vòng quay Tồn kho — {start.strftime('%d/%m/%Y')} đến {end.strftime('%d/%m/%Y')}",
    )


def _render_dashboard_summary_excel(data: Dict) -> Response:
    """Dashboard summary có 4 phần — gộp vào 1 sheet với section."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500, detail="openpyxl chưa được cài",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    section_font = Font(bold=True, size=11, color="1F4E79")

    row = 1
    ws.cell(row=row, column=1, value=f"Báo cáo Dashboard - tháng {data['month_label']}").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # Phần I: KPI
    ws.cell(row=row, column=1, value="I. Chỉ số tổng quan (KPI)").font = section_font
    row += 1
    kpi_headers = ["Chỉ số", "Giá trị", "Xu hướng"]
    for c, h in enumerate(kpi_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    kpi = data["kpi"]
    kpi_rows = [
        ("Tổng số ca hiện tại", kpi["total_cases_current"], f"{kpi['cases_trend_pct']:+.1f}%"),
        ("Số ca dự báo tháng tới", kpi["predicted_cases_next_month"], f"{kpi['predicted_trend_pct']:+.1f}%"),
        ("Vật tư thiếu hụt", f"{kpi['shortage_supplies_count']} mục", "—"),
        ("Mức nguy cơ chung", kpi["overall_risk"], "—"),
    ]
    for r in kpi_rows:
        for c, v in enumerate(r, 1):
            ws.cell(row=row, column=c, value=v)
        row += 1
    row += 1

    # Phần II: Xu hướng 6 tháng
    ws.cell(row=row, column=1, value="II. Xu hướng ca bệnh 6 tháng").font = section_font
    row += 1
    for c, h in enumerate(["Tháng", "Năm nay", "Năm trước"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    for r in data["case_trend"]:
        ws.cell(row=row, column=1, value=r["month"])
        ws.cell(row=row, column=2, value=r["this_year"])
        ws.cell(row=row, column=3, value=r["last_year"])
        row += 1
    row += 1

    # Phần III: Demand vs Stock
    ws.cell(row=row, column=1, value="III. Nhu cầu vs Tồn kho (Top 5)").font = section_font
    row += 1
    for c, h in enumerate(["Vật tư", "ĐVT", "Tồn kho", "Nhu cầu"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    for r in data["demand_vs_stock"]:
        ws.cell(row=row, column=1, value=r["supply_name"])
        ws.cell(row=row, column=2, value=r["unit"] or "")
        ws.cell(row=row, column=3, value=r["stock"])
        ws.cell(row=row, column=4, value=r["demand"])
        row += 1
    row += 1

    # Phần IV: Cảnh báo
    ws.cell(row=row, column=1, value="IV. Cảnh báo thiếu hụt vật tư (Top 5)").font = section_font
    row += 1
    for c, h in enumerate(["Vật tư", "Tồn hiện tại", "Định mức", "Trạng thái"], 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    severity_label = {"critical": "Nguy hiểm", "high": "Cần nhập", "medium": "Cảnh báo"}
    for a in data["alerts"]:
        ws.cell(row=row, column=1, value=a["supply_name"])
        ws.cell(row=row, column=2, value=a["current_stock"])
        ws.cell(row=row, column=3, value=a["required_stock"])
        ws.cell(row=row, column=4, value=severity_label.get(a["severity"], a["severity"]))
        row += 1

    for col_num, w in enumerate([34, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col_num)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"dashboard_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
